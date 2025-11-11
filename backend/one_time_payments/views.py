from decimal import Decimal
from collections import defaultdict

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View, generic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from stimuli.models import StimulusRequest, Employee
from stimuli.views import SortingMixin, resolve_sorting
from stimuli.forms import StimulusRequestStatusForm
from stimuli.filters import CampaignStimulusRequestFilter
from stimuli.services import recompute_employee_totals
from staffing.models import Division

from .forms import OneTimePaymentForm, RequestCampaignForm, RequestCampaignStatusForm
from .models import OneTimePayment, RequestCampaign


class RequestCampaignListView(LoginRequiredMixin, PermissionRequiredMixin, generic.ListView):
    model = RequestCampaign
    template_name = 'one_time_payments/campaign_list.html'
    context_object_name = 'campaigns'
    paginate_by = 25
    permission_required = 'one_time_payments.view_requestcampaign'

    def get_queryset(self):
        queryset = RequestCampaign.objects.all().order_by('-opens_at', 'name')
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = RequestCampaign.Status.choices
        context['selected_status'] = self.request.GET.get('status', '')
        return context


class RequestCampaignCreateView(LoginRequiredMixin, PermissionRequiredMixin, generic.CreateView):
    model = RequestCampaign
    form_class = RequestCampaignForm
    template_name = 'one_time_payments/campaign_form.html'
    permission_required = 'one_time_payments.add_requestcampaign'
    success_url = reverse_lazy('one_time_payments:campaign-list')

    def form_valid(self, form):
        messages.success(self.request, 'Кампания создана.')
        return super().form_valid(form)


class RequestCampaignUpdateView(LoginRequiredMixin, PermissionRequiredMixin, generic.UpdateView):
    model = RequestCampaign
    form_class = RequestCampaignForm
    template_name = 'one_time_payments/campaign_form.html'
    permission_required = 'one_time_payments.change_requestcampaign'

    def form_valid(self, form):
        messages.success(self.request, 'Кампания обновлена.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('one_time_payments:campaign-detail', args=[self.object.pk])


def aggregate_approved_requests(
    campaign,
    *,
    base_queryset=None,
    employee_ids=None,
    division_ids=None,
    responsible_ids=None
):
    qs = base_queryset
    if qs is None:
        qs = StimulusRequest.objects.filter(campaign=campaign)
    qs = qs.select_related('employee', 'employee__division', 'employee__position', 'requested_by')
    approved_qs = qs.filter(
        Q(status=StimulusRequest.Status.APPROVED) |
        Q(status=StimulusRequest.Status.ARCHIVED, final_status__icontains='Одобрено')
    )
    if employee_ids:
        approved_qs = approved_qs.filter(employee_id__in=employee_ids)
    if division_ids:
        approved_qs = approved_qs.filter(employee__division_id__in=division_ids)
    if responsible_ids:
        approved_qs = approved_qs.filter(requested_by_id__in=responsible_ids)

    grouped_requests = defaultdict(lambda: {
        'employee': None,
        'total_amount': Decimal('0'),
        'justifications': [],
        'amounts': [],
        'requesters': set(),
        'comments': [],
        'earliest_created_at': None,
    })

    for request in approved_qs:
        employee_id = request.employee.id
        grouped_requests[employee_id]['employee'] = request.employee
        grouped_requests[employee_id]['total_amount'] += request.amount

        if request.justification:
            grouped_requests[employee_id]['justifications'].append(request.justification)
            grouped_requests[employee_id]['amounts'].append(request.amount)

        requester_name = request.requested_by.get_full_name() or request.requested_by.username
        grouped_requests[employee_id]['requesters'].add(requester_name)

        if request.admin_comment:
            grouped_requests[employee_id]['comments'].append(request.admin_comment)

        earliest = grouped_requests[employee_id]['earliest_created_at']
        if earliest is None or request.created_at < earliest:
            grouped_requests[employee_id]['earliest_created_at'] = request.created_at

    approved_requests = []
    for data in grouped_requests.values():
        if len(data['justifications']) > 1:
            justifications_display = '; '.join([
                f"{just} ({amt} ₽)"
                for just, amt in zip(data['justifications'], data['amounts'])
            ])
        else:
            justifications_display = data['justifications'][0] if data['justifications'] else ''

        approved_requests.append({
            'employee': data['employee'],
            'total_amount': data['total_amount'],
            'justification': justifications_display,
            'requesters': ', '.join(sorted(data['requesters'])),
            'admin_comment': '; '.join(data['comments']) if data['comments'] else '',
            'created_at': data['earliest_created_at'],
        })

    approved_requests.sort(key=lambda x: x['employee'].full_name)

    return approved_requests


class RequestCampaignDetailView(SortingMixin, LoginRequiredMixin, PermissionRequiredMixin, generic.DetailView):
    model = RequestCampaign
    template_name = 'one_time_payments/campaign_detail.html'
    context_object_name = 'campaign'
    permission_required = 'one_time_payments.view_requestcampaign'

    SORTABLE_FIELDS = {
        'created': ('created_at',),
        'employee': ('employee__full_name',),
        'amount': ('amount', 'employee__full_name'),
        'status': ('status', 'employee__full_name'),
        'responsible': ('requested_by__last_name', 'requested_by__first_name', 'requested_by__username', 'employee__full_name'),
    }
    DEFAULT_SORT_FIELD = 'employee'
    DEFAULT_SORT_DIRECTION = 'asc'

    def get_queryset(self):
        sort_field, sort_direction, ordering = self._get_sorting_params()
        self.sort_field = sort_field
        self.sort_direction = sort_direction

        return RequestCampaign.objects.prefetch_related(
            Prefetch(
                'stimulus_requests',
                queryset=StimulusRequest.objects.select_related(
                    'employee', 'employee__division', 'employee__position', 'requested_by'
                ).order_by(*ordering),
            ),
            Prefetch(
                'manual_payments',
                queryset=OneTimePayment.objects.select_related('employee', 'created_by').order_by('-payment_date'),
            ),
        )

    def _build_query(self, *, exclude=None, overrides=None):
        params = self.request.GET.copy()
        exclude = exclude or []
        overrides = overrides or {}

        for key in exclude:
            params.pop(key, None)

        for key, value in overrides.items():
            if value is None:
                params.pop(key, None)
            else:
                params[key] = value

        encoded = params.urlencode()
        path = self.request.path
        return f'{path}?{encoded}' if encoded else path

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        campaign = self.object

        # Формируем список доступных действий в зависимости от статуса кампании
        available_actions = []
        if campaign.status == RequestCampaign.Status.DRAFT:
            available_actions = [('open', 'Открыть')]
        elif campaign.status == RequestCampaign.Status.OPEN:
            available_actions = [('close', 'Закрыть')]
        elif campaign.status == RequestCampaign.Status.CLOSED:
            available_actions = [('reopen', 'Переоткрыть'), ('archive', 'Переместить в архив')]
        # Для ARCHIVED действий нет

        context['status_form'] = RequestCampaignStatusForm()
        context['available_actions'] = available_actions
        context['stimulus_status_form'] = StimulusRequestStatusForm()

        base_requests_qs = StimulusRequest.objects.filter(
            campaign=campaign
        ).select_related(
            'employee',
            'employee__division',
            'employee__position',
            'requested_by',
        )

        params = self.request.GET.copy()
        for key in ('requested_by',):
            values = [value for value in params.getlist(key) if value != '__all__']
            if values:
                params.setlist(key, values)
            else:
                params.pop(key, None)
        
        # Убираем status из params для filterset - обработаем вручную
        params.pop('status', None)

        request_filter = CampaignStimulusRequestFilter(params or None, queryset=base_requests_qs)
        filtered_requests = request_filter.qs

        self.sort_field, self.sort_direction, ordering = self._get_sorting_params()
        filtered_requests = filtered_requests.order_by(*ordering)

        employee_ids = [
            value for value in base_requests_qs.values_list('employee_id', flat=True).distinct()
            if value is not None
        ]
        division_ids = [
            value for value in base_requests_qs.values_list('employee__division_id', flat=True).distinct()
            if value is not None
        ]
        employee_options = Employee.objects.filter(id__in=employee_ids).order_by('full_name')
        division_options = Division.objects.filter(id__in=division_ids).order_by('name')

        raw_employee_values = self.request.GET.getlist('employees')
        selected_employee_ids = []
        for value in raw_employee_values:
            try:
                selected_employee_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        if selected_employee_ids:
            filtered_requests = filtered_requests.filter(employee_id__in=selected_employee_ids)

        raw_division_values = self.request.GET.getlist('divisions')
        selected_division_ids = []
        for value in raw_division_values:
            try:
                selected_division_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        if selected_division_ids:
            filtered_requests = filtered_requests.filter(employee__division_id__in=selected_division_ids)

        # Обработка фильтра по статусу
        status_choices = list(StimulusRequest.Status.choices)
        valid_status_values = {choice[0] for choice in status_choices}
        raw_status_values = self.request.GET.getlist('status')
        selected_statuses = []
        for value in raw_status_values:
            if value == '__all__':
                continue
            if value in valid_status_values:
                selected_statuses.append(value)
        
        # Применяем фильтр по статусу к queryset
        if selected_statuses:
            filtered_requests = filtered_requests.filter(status__in=selected_statuses)
        
        context['filter'] = request_filter
        context['filter_form'] = request_filter.form
        context['requests'] = filtered_requests
        context['pending_requests_count'] = base_requests_qs.filter(status=StimulusRequest.Status.PENDING).count()
        context['employee_options'] = employee_options
        context['division_options'] = division_options
        context['selected_employee_ids'] = selected_employee_ids
        context['selected_division_ids'] = selected_division_ids
        context['status_options'] = status_choices
        context['selected_statuses'] = selected_statuses

        responsible_values_raw = self.request.GET.getlist('requested_by')
        selected_responsible_ids = []
        for value in responsible_values_raw:
            if value == '__all__':
                continue
            try:
                selected_responsible_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        context['selected_responsible_ids'] = selected_responsible_ids

        approved_employee_values = self.request.GET.getlist('approved_employees')
        approved_employee_ids = []
        for value in approved_employee_values:
            if value == '__all__':
                continue
            try:
                approved_employee_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        approved_division_values = self.request.GET.getlist('approved_divisions')
        approved_division_ids = []
        for value in approved_division_values:
            if value == '__all__':
                continue
            try:
                approved_division_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        approved_responsible_values = self.request.GET.getlist('approved_responsible')
        approved_responsible_ids = []
        for value in approved_responsible_values:
            if value == '__all__':
                continue
            try:
                approved_responsible_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        approved_base_qs = base_requests_qs
        if approved_employee_ids:
            approved_base_qs = approved_base_qs.filter(employee_id__in=approved_employee_ids)
        if approved_division_ids:
            approved_base_qs = approved_base_qs.filter(employee__division_id__in=approved_division_ids)
        if approved_responsible_ids:
            approved_base_qs = approved_base_qs.filter(requested_by_id__in=approved_responsible_ids)

        approved_requests = aggregate_approved_requests(
            campaign,
            base_queryset=approved_base_qs,
            employee_ids=approved_employee_ids,
            division_ids=approved_division_ids,
            responsible_ids=approved_responsible_ids,
        )

        context['approved_requests'] = approved_requests
        
        # Формируем списки опций для фильтров одобренных заявок (только из одобренных)
        approved_only_qs = base_requests_qs.filter(
            Q(status=StimulusRequest.Status.APPROVED) |
            Q(status=StimulusRequest.Status.ARCHIVED, final_status__icontains='Одобрено')
        )
        
        # Опции сотрудников для фильтра одобренных заявок
        approved_employee_ids_for_options = [
            value for value in approved_only_qs.values_list('employee_id', flat=True).distinct()
            if value is not None
        ]
        approved_employee_options = Employee.objects.filter(
            id__in=approved_employee_ids_for_options
        ).order_by('full_name')
        
        # Опции подразделений для фильтра одобренных заявок
        approved_division_ids_for_options = [
            value for value in approved_only_qs.values_list('employee__division_id', flat=True).distinct()
            if value is not None
        ]
        approved_division_options = Division.objects.filter(
            id__in=approved_division_ids_for_options
        ).order_by('name')
        
        # Опции ответственных для фильтра одобренных заявок
        responsible_qs = approved_only_qs.values_list(
            'requested_by_id',
            'requested_by__first_name',
            'requested_by__last_name',
            'requested_by__username',
        ).distinct()
        responsible_map = {}
        for user_id, first_name, last_name, username in responsible_qs:
            full_name = ' '.join(filter(None, [last_name, first_name])).strip()
            display = full_name or username or f'ID {user_id}'
            if user_id is not None:
                responsible_map[user_id] = display

        approved_responsible_options = [
            {'id': user_id, 'display': responsible_map[user_id]}
            for user_id in sorted(responsible_map.keys(), key=lambda pk: responsible_map[pk].lower())
        ]
        
        # Опции ответственных для фильтра всех заявок
        all_responsible_qs = base_requests_qs.values_list(
            'requested_by_id',
            'requested_by__first_name',
            'requested_by__last_name',
            'requested_by__username',
        ).distinct()
        all_responsible_map = {}
        for user_id, first_name, last_name, username in all_responsible_qs:
            full_name = ' '.join(filter(None, [last_name, first_name])).strip()
            display = full_name or username or f'ID {user_id}'
            if user_id is not None:
                all_responsible_map[user_id] = display

        responsible_options = [
            {'id': user_id, 'display': all_responsible_map[user_id]}
            for user_id in sorted(all_responsible_map.keys(), key=lambda pk: all_responsible_map[pk].lower())
        ]

        context['approved_filters'] = {
            'employees': approved_employee_ids,
            'divisions': approved_division_ids,
            'responsibles': approved_responsible_ids,
        }
        context['approved_employee_options'] = approved_employee_options
        context['approved_division_options'] = approved_division_options
        context['approved_responsible_options'] = approved_responsible_options
        context['approved_employee_ids'] = approved_employee_ids
        context['approved_division_ids'] = approved_division_ids
        context['approved_responsible_ids'] = approved_responsible_ids
        context['responsible_options'] = responsible_options

        context['sorting'] = self._build_sorting_context()
        context['requests_reset_url'] = self._build_query(exclude=['employees', 'divisions', 'status', 'requested_by'])
        context['approved_reset_url'] = self._build_query(exclude=['approved_employees', 'approved_divisions', 'approved_responsible'])
        
        # Добавляем сводку по запрошенным средствам
        context['amounts_summary'] = campaign.get_requested_amounts_summary()

        return context


class RequestCampaignStatusUpdateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'one_time_payments.change_requestcampaign'

    def post(self, request, *args, **kwargs):
        campaign = get_object_or_404(RequestCampaign, pk=kwargs['pk'])
        form = RequestCampaignStatusForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'Не удалось обновить статус кампании.')
            return redirect('one_time_payments:campaign-detail', pk=campaign.pk)

        action = form.cleaned_data['action']
        try:
            if action == 'open':
                campaign.open()
                messages.success(request, 'Кампания открыта.')
            elif action == 'close':
                campaign.close(archive=False)
                messages.success(request, 'Кампания закрыта.')
            elif action == 'reopen':
                campaign.reopen()
                messages.success(request, 'Кампания переоткрыта.')
            elif action == 'archive':
                campaign.archive()
                messages.success(request, 'Кампания архивирована.')
        except ValidationError as exc:
            messages.error(request, exc.message)
        return redirect('one_time_payments:campaign-detail', pk=campaign.pk)


class ManualStimulusStatusUpdateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'stimuli.change_stimulusrequest'

    def post(self, request, *args, **kwargs):
        campaign = get_object_or_404(RequestCampaign, pk=kwargs['pk'])
        stimulus = get_object_or_404(StimulusRequest, pk=kwargs['request_pk'], campaign=campaign)
        form = StimulusRequestStatusForm(request.POST, instance=stimulus)
        if form.is_valid():
            previous_employee_id = stimulus.employee_id
            updated_request = form.save()
            recompute_employee_totals(updated_request.employee_id)
            if previous_employee_id != updated_request.employee_id:
                recompute_employee_totals(previous_employee_id)
            messages.success(request, 'Статус заявки обновлён.')
        else:
            messages.error(request, 'Не удалось обновить заявку.')
        return redirect('one_time_payments:campaign-detail', pk=campaign.pk)


class CampaignApprovePendingRequestsView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'stimuli.change_stimulusrequest'

    def post(self, request, *args, **kwargs):
        campaign = get_object_or_404(RequestCampaign, pk=kwargs['pk'])
        pending_requests = StimulusRequest.objects.filter(
            campaign=campaign,
            status=StimulusRequest.Status.PENDING,
        ).select_related('employee')

        if not pending_requests.exists():
            messages.info(request, 'Нет заявок на рассмотрении для одобрения.')
            return redirect('one_time_payments:campaign-detail', pk=campaign.pk)

        updated_count = 0
        affected_employee_ids = set()

        with transaction.atomic():
            for stimulus in pending_requests:
                if stimulus.status != StimulusRequest.Status.PENDING:
                    continue
                stimulus.status = StimulusRequest.Status.APPROVED
                stimulus.save(update_fields=['status', 'updated_at'])
                affected_employee_ids.add(stimulus.employee_id)
                updated_count += 1

        for employee_id in affected_employee_ids:
            recompute_employee_totals(employee_id)

        messages.success(request, f'Одобрено заявок: {updated_count}.')
        return redirect('one_time_payments:campaign-detail', pk=campaign.pk)


class ManualPaymentCreateView(LoginRequiredMixin, PermissionRequiredMixin, generic.CreateView):
    model = OneTimePayment
    form_class = OneTimePaymentForm
    template_name = 'one_time_payments/manual_payment_form.html'
    permission_required = 'one_time_payments.add_onetimepayment'

    def get_initial(self):
        initial = super().get_initial()
        campaign_id = self.kwargs.get('pk') or self.request.GET.get('campaign')
        if campaign_id:
            try:
                initial['campaign'] = RequestCampaign.objects.get(pk=campaign_id)
            except (RequestCampaign.DoesNotExist, ValueError, TypeError):
                pass
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Разовая выплата создана.')
        return super().form_valid(form)

    def form_invalid(self, form):
        campaign_id = form.data.get('campaign')
        error_text = '; '.join([f"{field}: {', '.join(errors)}" for field, errors in form.errors.items()]) or 'Проверьте введённые данные.'
        messages.error(self.request, f'Не удалось создать выплату. {error_text}')
        if campaign_id:
            return redirect('one_time_payments:campaign-detail', pk=campaign_id)
        return self.render_to_response(self.get_context_data(form=form))

    def get_success_url(self):
        if self.object.campaign_id:
            return reverse('one_time_payments:campaign-detail', args=[self.object.campaign_id])
        return reverse('one_time_payments:manual-payment-list')


class ManualPaymentListView(LoginRequiredMixin, PermissionRequiredMixin, generic.ListView):
    model = OneTimePayment
    template_name = 'one_time_payments/manual_payment_list.html'
    context_object_name = 'payments'
    paginate_by = 25
    permission_required = 'one_time_payments.view_onetimepayment'

    def get_queryset(self):
        queryset = OneTimePayment.objects.select_related('employee', 'campaign', 'created_by').order_by('-payment_date', '-created_at')
        campaign_id = self.request.GET.get('campaign')
        if campaign_id:
            queryset = queryset.filter(campaign_id=campaign_id)
            self.campaign_id = campaign_id
            self.campaign_obj = RequestCampaign.objects.filter(pk=campaign_id).first()
        else:
            self.campaign_id = None
            self.campaign_obj = None
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['selected_campaign'] = getattr(self, 'campaign_id', None)
        context['selected_campaign_obj'] = getattr(self, 'campaign_obj', None)
        context['campaigns'] = RequestCampaign.objects.order_by('-opens_at', 'name')
        return context


class ManualPaymentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, generic.UpdateView):
    model = OneTimePayment
    form_class = OneTimePaymentForm
    template_name = 'one_time_payments/manual_payment_form.html'
    permission_required = 'one_time_payments.change_onetimepayment'

    def form_valid(self, form):
        messages.success(self.request, 'Разовая выплата обновлена.')
        return super().form_valid(form)

    def get_success_url(self):
        if self.object.campaign_id:
            return reverse('one_time_payments:campaign-detail', args=[self.object.campaign_id])
        return reverse('one_time_payments:manual-payment-list')


class ManualPaymentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, generic.DeleteView):
    model = OneTimePayment
    template_name = 'one_time_payments/manual_payment_confirm_delete.html'
    permission_required = 'one_time_payments.delete_onetimepayment'

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        campaign_id = self.object.campaign_id
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'Разовая выплата удалена.')
        if campaign_id:
            return redirect('one_time_payments:campaign-detail', pk=campaign_id)
        return redirect('one_time_payments:manual-payment-list')

    def get_success_url(self):
        if self.object.campaign_id:
            return reverse('one_time_payments:campaign-detail', args=[self.object.campaign_id])
        return reverse('one_time_payments:manual-payment-list')


class CampaignApprovedRequestsExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Экспорт одобренных заявок кампании в Excel"""
    permission_required = 'one_time_payments.view_requestcampaign'

    def get(self, request, *args, **kwargs):
        campaign = get_object_or_404(
            RequestCampaign,
            pk=kwargs['pk']
        )
        
        # Получаем одобренные заявки кампании
        # Включаем как текущие одобренные, так и одобренные до архивирования
        approved_employee_values = request.GET.getlist('approved_employees')
        approved_employee_ids = []
        for value in approved_employee_values:
            if value == '__all__':
                continue
            try:
                approved_employee_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        approved_division_values = request.GET.getlist('approved_divisions')
        approved_division_ids = []
        for value in approved_division_values:
            if value == '__all__':
                continue
            try:
                approved_division_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        approved_responsible_values = request.GET.getlist('approved_responsible')
        approved_responsible_ids = []
        for value in approved_responsible_values:
            if value == '__all__':
                continue
            try:
                approved_responsible_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        approved_base_qs = StimulusRequest.objects.filter(campaign=campaign)
        if approved_employee_ids:
            approved_base_qs = approved_base_qs.filter(employee_id__in=approved_employee_ids)
        if approved_division_ids:
            approved_base_qs = approved_base_qs.filter(employee__division_id__in=approved_division_ids)
        if approved_responsible_ids:
            approved_base_qs = approved_base_qs.filter(requested_by_id__in=approved_responsible_ids)

        approved_requests_grouped = aggregate_approved_requests(
            campaign,
            base_queryset=approved_base_qs,
            employee_ids=approved_employee_ids,
            division_ids=approved_division_ids,
            responsible_ids=approved_responsible_ids,
        )

        # Создаем Excel файл
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Одобренные заявки'
        
        # Заголовки
        headers = [
            'ФИО сотрудника',
            'Подразделение',
            'Должность',
            'Итоговая сумма',
            'Обоснование',
            'Ответственные',
            'Комментарий',
            'Дата создания'
        ]
        sheet.append(headers)
        
        # Стили для заголовков
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Данные
        for item in approved_requests_grouped:
            sheet.append([
                item['employee'].full_name,
                item['employee'].division.name if item['employee'].division else '',
                item['employee'].position.name if item['employee'].position else '',
                float(item['total_amount']),
                item['justification'],
                item['requesters'],
                item['admin_comment'] or '',
                item['created_at'].strftime('%d.%m.%Y %H:%M')
            ])
        
        # Автоматическая ширина столбцов
        for idx, column in enumerate(sheet.columns, start=1):
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0 
                for cell in column
            )
            adjusted_width = max(10, min(max_length + 2, 50))
            sheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = 'A2'

        UserModel = get_user_model()
        if approved_responsible_ids:
            responsible_names = [
                user.get_full_name() or user.username
                for user in UserModel.objects.filter(id__in=approved_responsible_ids).order_by('last_name', 'first_name', 'username')
            ]
        else:
            responsible_names = []

        if approved_employee_ids or approved_division_ids or approved_responsible_ids:
            filters_sheet = workbook.create_sheet('Фильтры')
            headers = ['Поле', 'Значение']
            filters_sheet.append(headers)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for cell in filters_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if approved_employee_ids:
                employee_names = list(
                    Employee.objects.filter(id__in=approved_employee_ids).order_by('full_name').values_list('full_name', flat=True)
                )
                filters_sheet.append(['Сотрудники', ', '.join(employee_names)])
            if approved_division_ids:
                division_names = list(
                    Division.objects.filter(id__in=approved_division_ids).order_by('name').values_list('name', flat=True)
                )
                filters_sheet.append(['Подразделения', ', '.join(division_names)])
            if responsible_names:
                filters_sheet.append(['Ответственные', ', '.join(responsible_names)])

            for column_cells in filters_sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                column_letter = get_column_letter(column_cells[0].column)
                filters_sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Сохраняем файл в ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = timezone.now().strftime('%Y%m%d_%H%M')
        campaign_name_clean = "".join(c for c in campaign.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        response['Content-Disposition'] = f'attachment; filename="campaign_{campaign_name_clean}_{timestamp}.xlsx"'
        
        workbook.save(response)
        return response


class CampaignRequestsExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Экспорт всех заявок кампании в Excel."""
    permission_required = 'one_time_payments.view_requestcampaign'

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    def get(self, request, *args, **kwargs):
        campaign = get_object_or_404(RequestCampaign, pk=kwargs['pk'])
        base_qs = StimulusRequest.objects.filter(campaign=campaign).select_related(
            'employee',
            'employee__division',
            'employee__position',
            'requested_by',
        )

        params = request.GET.copy()
        for key in ('status', 'requested_by'):
            values = [value for value in params.getlist(key) if value != '__all__']
            if values:
                params.setlist(key, values)
            else:
                params.pop(key, None)

        filterset = CampaignStimulusRequestFilter(params or None, queryset=base_qs)
        filtered_qs = filterset.qs
        employee_values = request.GET.getlist('employees')
        employee_filter_ids = []
        for value in employee_values:
            try:
                employee_filter_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        if employee_filter_ids:
            filtered_qs = filtered_qs.filter(employee_id__in=employee_filter_ids)

        division_values = request.GET.getlist('divisions')
        division_filter_ids = []
        for value in division_values:
            try:
                division_filter_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        if division_filter_ids:
            filtered_qs = filtered_qs.filter(employee__division_id__in=division_filter_ids)
        status_map = dict(StimulusRequest.Status.choices)
        status_filter_values = [
            value for value in params.getlist('status') if value in status_map
        ]
        responsible_filter_ids = []
        for value in params.getlist('requested_by'):
            try:
                responsible_filter_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        _, _, ordering = resolve_sorting(
            request,
            RequestCampaignDetailView.SORTABLE_FIELDS,
            RequestCampaignDetailView.DEFAULT_SORT_FIELD,
            RequestCampaignDetailView.DEFAULT_SORT_DIRECTION,
        )
        stimulus_requests = filtered_qs.order_by(*ordering)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Заявки кампании'

        headers = [
            'ID',
            'Создано',
            'Сотрудник',
            'Подразделение',
            'Должность',
            'Размер выплаты',
            'Статус',
            'Итоговый статус',
            'Ответственный',
            'Обоснование',
            'Комментарий администратора',
            'Дата архивации',
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment

        for stimulus in stimulus_requests:
            sheet.append([
                stimulus.pk,
                stimulus.created_at.strftime('%d.%m.%Y %H:%M') if stimulus.created_at else '',
                stimulus.employee.full_name,
                stimulus.employee.division.name if stimulus.employee and stimulus.employee.division else '',
                stimulus.employee.position.name if stimulus.employee and stimulus.employee.position else '',
                float(stimulus.amount) if stimulus.amount is not None else '',
                stimulus.get_status_display(),
                stimulus.final_status or '',
                stimulus.requested_by.get_full_name() or stimulus.requested_by.username,
                stimulus.justification,
                stimulus.admin_comment,
                stimulus.archived_at.strftime('%d.%m.%Y %H:%M') if stimulus.archived_at else '',
            ])

        self._autosize_columns(sheet)
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = 'A2'

        extra_filter_rows = []
        if employee_filter_ids:
            employee_names = list(
                Employee.objects.filter(id__in=employee_filter_ids).order_by('full_name').values_list('full_name', flat=True)
            )
            extra_filter_rows.append(('Сотрудники', ', '.join(employee_names)))
        if division_filter_ids:
            division_names = list(
                Division.objects.filter(id__in=division_filter_ids).order_by('name').values_list('name', flat=True)
            )
            extra_filter_rows.append(('Подразделения', ', '.join(division_names)))

        if status_filter_values:
            extra_filter_rows.append(('Статусы', ', '.join(status_map[value] for value in status_filter_values)))
        if responsible_filter_ids:
            UserModel = get_user_model()
            responsible_names = [
                user.get_full_name() or user.username
                for user in UserModel.objects.filter(id__in=responsible_filter_ids).order_by('last_name', 'first_name', 'username')
            ]
            extra_filter_rows.append(('Ответственные', ', '.join(responsible_names)))

        if filterset is not None and params:
            filters_sheet = workbook.create_sheet("Фильтры")
            self._write_filters_sheet(filters_sheet, filterset)
            if extra_filter_rows:
                filters_sheet.append(['', ''])
                for label, value in extra_filter_rows:
                    filters_sheet.append([label, value])
                self._autosize_columns(filters_sheet)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = timezone.now().strftime('%Y%m%d_%H%M')
        campaign_name_clean = "".join(
            c for c in campaign.name if c.isalnum() or c in (' ', '-', '_')
        ).strip()
        response['Content-Disposition'] = (
            f'attachment; filename="campaign_{campaign_name_clean}_requests_{timestamp}.xlsx"'
        )

        workbook.save(response)
        return response

    def _autosize_columns(self, worksheet):
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                try:
                    length = len(str(value))
                except (TypeError, ValueError):
                    length = 0
                if length > max_length:
                    max_length = length
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
