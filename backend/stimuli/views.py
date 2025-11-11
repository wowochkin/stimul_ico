from decimal import Decimal
import io
import logging
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import transaction
from django.db.models import BooleanField, Case, Value, When, Q, QuerySet
from django.http import Http404, QueryDict, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.views import View, generic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .filters import EmployeeFilter, StimulusRequestFilter
from .forms import EmployeeForm, InternalAssignmentFormSet, StimulusRequestForm, StimulusRequestStatusForm, EmployeeExcelUploadForm
from .permissions import (
    is_department_manager, is_employee, get_user_division,
    can_view_all_requests, can_view_own_requests, can_edit_request, can_delete_request, can_change_request_status, get_accessible_employees
)
from one_time_payments.models import RequestCampaign
from staffing.models import Division, Position
from .models import Employee, StimulusRequest
from .services import recompute_employee_totals


def resolve_sorting(request, sortable_fields, default_field='', default_direction='asc'):
    sort_field = request.GET.get('sort') or default_field
    if sort_field not in sortable_fields:
        sort_field = default_field

    sort_direction = request.GET.get('direction') or default_direction
    if sort_direction not in ('asc', 'desc'):
        sort_direction = default_direction

    ordering = []
    for field in sortable_fields.get(sort_field, ()):
        ordering.append(f'-{field}' if sort_direction == 'desc' else field)
    ordering.append('-pk' if sort_direction == 'desc' else 'pk')

    return sort_field, sort_direction, ordering


class SortingMixin:
    """Mixin с общими помощниками для сортировки списков."""

    SORTABLE_FIELDS = {}
    DEFAULT_SORT_FIELD = ''
    DEFAULT_SORT_DIRECTION = 'asc'

    def _get_sorting_params(self):
        return resolve_sorting(
            self.request,
            self.SORTABLE_FIELDS,
            self.DEFAULT_SORT_FIELD,
            self.DEFAULT_SORT_DIRECTION,
        )

    def _build_sorting_context(self):
        base_params = self.request.GET.copy()
        for key in ('page', 'sort', 'direction'):
            base_params.pop(key, None)

        sorting = {}
        current_sort = getattr(self, 'sort_field', self.DEFAULT_SORT_FIELD)
        current_direction = getattr(self, 'sort_direction', self.DEFAULT_SORT_DIRECTION)

        for key in self.SORTABLE_FIELDS.keys():
            params = base_params.copy()
            is_active = key == current_sort
            direction_now = current_direction if is_active else None
            next_direction = 'desc' if is_active and current_direction == 'asc' else 'asc'
            params['sort'] = key
            params['direction'] = next_direction
            query_string = params.urlencode()
            url = f'?{query_string}' if query_string else f'?sort={key}&direction={next_direction}'
            sorting[key] = {
                'url': url,
                'is_active': is_active,
                'current_direction': direction_now,
                'next_direction': next_direction,
            }

        sorting['current_sort'] = current_sort
        sorting['current_direction'] = current_direction
        return sorting


class EmployeeListView(LoginRequiredMixin, PermissionRequiredMixin, generic.ListView):
    model = Employee
    template_name = 'stimuli/employee_list.html'
    context_object_name = 'employees'
    paginate_by = 25
    permission_required = 'stimuli.view_employee'
    AVAILABLE_COLUMNS = [
        ('division', 'Подразделение'),
        ('position', 'Должность'),
        ('base_salary', 'Оклад'),
        ('rate', 'Ставка'),
        ('salary', 'Выплаты по ставке'),
        ('assignments', 'Совмещения'),
        ('assignments_salary', 'Оклад совмещений'),
        ('total_salary', 'Итого базовых выплат'),
        ('allowance_amount', 'Надбавка'),
        ('allowance_reason', 'Основание надбавки'),
        ('allowance_until', 'Срок надбавки'),
        ('payment', 'Выплата'),
        ('justification', 'Обоснование'),
        ('total_payments', 'Итого выплат'),
    ]
    DEFAULT_COLUMNS = [
        'division',
        'position',
        'base_salary',
        'rate',
        'salary',
        'assignments',
        'assignments_salary',
        'total_salary',
        'allowance_amount',
        'allowance_reason',
        'allowance_until',
        'payment',
        'justification',
        'total_payments',
    ]

    def get_paginate_by(self, queryset):
        if self.request.GET.get('show') == 'all':
            return None
        return self.paginate_by

    def get_queryset(self):
        qs = Employee.objects.select_related('division', 'position').prefetch_related('requests__requested_by', 'assignments__position')
        self.filterset = EmployeeFilter(self.request.GET or None, queryset=qs)
        return self.filterset.qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = getattr(
            self,
            'filterset',
            EmployeeFilter(None, queryset=self.model.objects.all())
        )
        selected_columns = self.get_selected_columns()
        user = self.request.user
        show_all = self.request.GET.get('show') == 'all'
        context.update({
            'available_columns': self.AVAILABLE_COLUMNS,
            'selected_columns': selected_columns,
            'show_all': show_all,
            'show_all_url': self._build_query(show='all'),
            'paginate_url': self._build_query(show=None),
        })
        extra_cols = 1 + len(selected_columns)
        if user.has_perm('stimuli.change_employee'):
            extra_cols += 1
        context['table_colspan'] = extra_cols
        return context

    def get_selected_columns(self):
        valid_keys = [key for key, _ in self.AVAILABLE_COLUMNS]
        requested = self.request.GET.getlist('columns')
        if requested:
            selected = [key for key in requested if key in valid_keys]
        else:
            selected = list(self.DEFAULT_COLUMNS)
        if not selected:
            selected = list(self.DEFAULT_COLUMNS)
        return selected

    def _build_query(self, **updates):
        params = self.request.GET.copy()
        for key, value in updates.items():
            if value is None:
                params.pop(key, None)
            else:
                params[key] = value
        encoded = params.urlencode()
        return '?' + encoded if encoded else ''


class EmployeeCreateView(LoginRequiredMixin, PermissionRequiredMixin, generic.CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'stimuli/employee_form.html'
    permission_required = 'stimuli.add_employee'
    success_url = reverse_lazy('employee-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['assignment_formset'] = InternalAssignmentFormSet(self.request.POST)
        else:
            context['assignment_formset'] = InternalAssignmentFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        assignment_formset = context['assignment_formset']
        if assignment_formset.is_valid():
            self.object = form.save()
            assignment_formset.instance = self.object
            assignment_formset.save()
            return redirect(self.success_url)
        return self.form_invalid(form)

    def form_invalid(self, form):
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)


class EmployeeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, generic.UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'stimuli/employee_form.html'
    permission_required = 'stimuli.change_employee'
    success_url = reverse_lazy('employee-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['assignment_formset'] = InternalAssignmentFormSet(self.request.POST, instance=self.object)
        else:
            context['assignment_formset'] = InternalAssignmentFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        assignment_formset = context['assignment_formset']
        if assignment_formset.is_valid():
            self.object = form.save()
            assignment_formset.instance = self.object
            assignment_formset.save()
            return redirect(self.success_url)
        return self.form_invalid(form)

    def form_invalid(self, form):
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)


class EmployeeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, generic.DeleteView):
    model = Employee
    template_name = 'stimuli/employee_confirm_delete.html'
    permission_required = 'stimuli.delete_employee'
    success_url = reverse_lazy('employee-list')


class StimulusRequestListView(SortingMixin, LoginRequiredMixin, generic.ListView):
    model = StimulusRequest
    template_name = 'stimuli/request_list.html'
    context_object_name = 'requests'
    paginate_by = 25

    SORTABLE_FIELDS = {
        'created': ('created_at',),
        'employee': ('employee__full_name',),
        'campaign': ('campaign__name', 'employee__full_name'),
        'amount': ('amount', 'employee__full_name'),
        'status': ('status', 'employee__full_name'),
        'responsible': ('requested_by__last_name', 'requested_by__first_name', 'requested_by__username', 'employee__full_name'),
    }
    DEFAULT_SORT_FIELD = 'employee'
    DEFAULT_SORT_DIRECTION = 'asc'

    def get_queryset(self):
        qs = StimulusRequest.objects.select_related('employee', 'requested_by', 'campaign')
        user = self.request.user
        
        # Определяем базовый queryset в зависимости от прав пользователя
        # ВАЖНО: Проверяем can_view_own_requests ПЕРВЫМ, чтобы он имел приоритет
        if can_view_own_requests(user):
            # Пользователи с can_view_own_requests видят:
            # 1. Заявки на самого себя (employee = user.employee_profile)
            # 2. Заявки, которые они сами подали (requested_by = user)
            try:
                employee = user.employee_profile
                base_qs = qs.filter(
                    Q(employee=employee) | Q(requested_by=user)
                )
            except Employee.DoesNotExist:
                # Если нет связи с Employee, показываем только заявки, которые пользователь сам подал
                base_qs = qs.filter(requested_by=user)
        elif can_view_all_requests(user):
            # Администраторы, руководство института и пользователи с can_view_all видят все заявки
            if user.is_staff or user.groups.filter(name='Руководство института').exists():
                base_qs = qs
            else:
                # Пользователи с can_view_all (но не администраторы и не руководство института) видят только свои заявки
                base_qs = qs.filter(requested_by=user)
        elif is_department_manager(user):
            user_division = get_user_division(user)
            if user_division:
                base_qs = qs.filter(employee__division=user_division)
            else:
                base_qs = qs.none()
        elif is_employee(user):
            base_qs = qs.filter(requested_by=user)
        else:
            base_qs = qs.none()
        
        # Обрабатываем параметры для filterset (только campaign и requested_by)
        params = self.request.GET.copy()
        for key in ('requested_by',):
            values = [value for value in params.getlist(key) if value != '__all__']
            if values:
                params.setlist(key, values)
            else:
                params.pop(key, None)
        
        # Убираем status из params для filterset - обработаем вручную
        params.pop('status', None)

        self.filterset = StimulusRequestFilter(params or None, queryset=base_qs, request=self.request)
        filtered_qs = self.filterset.qs

        base_for_options = self.filterset.queryset if hasattr(self.filterset, 'queryset') else base_qs
        employee_ids = [
            value for value in base_for_options.values_list('employee_id', flat=True).distinct()
            if value is not None
        ]
        division_ids = [
            value for value in base_for_options.values_list('employee__division_id', flat=True).distinct()
            if value is not None
        ]
        self.employee_options = list(Employee.objects.filter(id__in=employee_ids).order_by('full_name'))
        self.division_options = list(Division.objects.filter(id__in=division_ids).order_by('name'))

        raw_employee_values = self.request.GET.getlist('employees')
        selected_employee_ids = []
        for value in raw_employee_values:
            if value == '__all__':
                continue
            try:
                selected_employee_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        if selected_employee_ids:
            filtered_qs = filtered_qs.filter(employee_id__in=selected_employee_ids)

        raw_division_values = self.request.GET.getlist('divisions')
        selected_division_ids = []
        for value in raw_division_values:
            if value == '__all__':
                continue
            try:
                selected_division_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        if selected_division_ids:
            filtered_qs = filtered_qs.filter(employee__division_id__in=selected_division_ids)

        self.selected_employee_ids = selected_employee_ids
        self.selected_division_ids = selected_division_ids

        # Обработка фильтра по статусу
        status_choices = list(StimulusRequest.Status.choices)
        valid_status_values = {choice[0] for choice in status_choices}
        raw_status_values = self.request.GET.getlist('status')
        selected_statuses = []
        for value in raw_status_values:
            if value == '__all__':
                continue
            if value in valid_status_values:
                selected_statuses.append(value)
        
        # Применяем фильтр по статусу к queryset
        if selected_statuses:
            filtered_qs = filtered_qs.filter(status__in=selected_statuses)
        
        self.status_options = status_choices
        self.selected_statuses = selected_statuses

        responsible_queryset = self.filterset.form.fields['requested_by'].queryset
        self.responsible_options = list(responsible_queryset)
        raw_responsible_values = self.request.GET.getlist('requested_by')
        selected_responsible_ids = []
        for value in raw_responsible_values:
            if value == '__all__':
                continue
            try:
                selected_responsible_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        self.selected_responsible_ids = selected_responsible_ids

        # Обработка фильтра кампаний
        from one_time_payments.models import RequestCampaign
        if user.is_staff or user.groups.filter(name='Руководство института').exists():
            # Администраторы и руководство видят все кампании кроме черновиков
            campaign_queryset = RequestCampaign.objects.exclude(
                status=RequestCampaign.Status.DRAFT
            ).order_by('-opens_at', 'name')
        else:
            # Остальные видят только открытые кампании
            campaign_queryset = RequestCampaign.objects.filter(
                status=RequestCampaign.Status.OPEN
            ).order_by('-opens_at', 'name')
        
        # Получаем все кампании, которые есть в отфильтрованных заявках
        campaign_ids_in_requests = base_for_options.values_list('campaign_id', flat=True).distinct()
        self.campaign_options = list(campaign_queryset.filter(id__in=campaign_ids_in_requests))
        
        raw_campaign_values = self.request.GET.getlist('campaign')
        selected_campaign_ids = []
        for value in raw_campaign_values:
            if value == '__all__':
                continue
            try:
                selected_campaign_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        if selected_campaign_ids:
            filtered_qs = filtered_qs.filter(campaign_id__in=selected_campaign_ids)
        
        self.selected_campaign_ids = selected_campaign_ids

        self.sort_field, self.sort_direction, ordering = self._get_sorting_params()
        ordered_qs = filtered_qs.order_by(*ordering)
        
        # Добавляем аннотации для определения прав редактирования и удаления
        if can_view_all_requests(user) and user.is_staff:
            # Администраторы имеют полный доступ
            return ordered_qs.annotate(
                can_edit=Value(True, output_field=BooleanField()),
                can_delete=Value(True, output_field=BooleanField()),
                can_change_status=Value(True, output_field=BooleanField()),
            )
        elif can_view_all_requests(user) and user.groups.filter(name='Руководство института').exists():
            # Руководство института видит все заявки, но редактировать/удалять может только свои в статусе PENDING
            return ordered_qs.annotate(
                can_edit=Case(
                    When(
                        requested_by=user,
                        status=StimulusRequest.Status.PENDING,
                        then=Value(True)
                    ),
                    default=Value(False),
                    output_field=BooleanField()
                ),
                can_delete=Case(
                    When(
                        requested_by=user,
                        status=StimulusRequest.Status.PENDING,
                        then=Value(True)
                    ),
                    default=Value(False),
                    output_field=BooleanField()
                ),
                can_change_status=Value(False, output_field=BooleanField()),
            )

        # Для остальных пользователей проверяем права для каждой заявки
        annotated_qs = ordered_qs.annotate(
            can_edit=Value(False, output_field=BooleanField()),
            can_delete=Value(False, output_field=BooleanField()),
            can_change_status=Value(False, output_field=BooleanField()),
        )
        
        # Обновляем аннотации для каждой заявки
        for request_obj in annotated_qs:
            request_obj.can_edit = can_edit_request(user, request_obj)
            request_obj.can_delete = can_delete_request(user, request_obj)
            request_obj.can_change_status = can_change_request_status(user, request_obj)
        
        return annotated_qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filterset
        context['filter_form'] = self.filterset.form
        user = self.request.user
        
        # Определяем права на основе роли пользователя
        can_bulk_delete = (
            can_view_all_requests(user) or 
            (is_department_manager(user) and get_user_division(user)) or
            (is_employee(user) and user.has_perm('stimuli.edit_pending_requests'))
        )
        
        show_manage = (
            can_view_all_requests(user) or 
            (is_department_manager(user) and get_user_division(user)) or
            (is_employee(user) and user.has_perm('stimuli.edit_pending_requests'))
        )
        
        context['can_bulk_delete'] = can_bulk_delete
        context['show_manage_column'] = show_manage
        context['is_department_manager'] = is_department_manager(user)
        context['is_employee'] = is_employee(user)
        context['user_division'] = get_user_division(user)
        context['filter_form'] = self.filterset.form
        export_url = reverse('request-export')
        query_string = self.request.GET.urlencode()
        if query_string:
            export_url = f'{export_url}?{query_string}'
        context['export_url'] = export_url
        
        base_columns = 8
        if can_bulk_delete:
            base_columns += 1
        if show_manage:
            base_columns += 1
        context['table_colspan'] = base_columns
        context['sorting'] = self._build_sorting_context()
        context['employee_options'] = getattr(self, 'employee_options', [])
        context['division_options'] = getattr(self, 'division_options', [])
        context['selected_employee_ids'] = getattr(self, 'selected_employee_ids', [])
        context['selected_division_ids'] = getattr(self, 'selected_division_ids', [])
        context['status_options'] = getattr(self, 'status_options', [])
        context['selected_statuses'] = getattr(self, 'selected_statuses', [])
        context['responsible_options'] = getattr(self, 'responsible_options', [])
        context['selected_responsible_ids'] = getattr(self, 'selected_responsible_ids', [])
        context['campaign_options'] = getattr(self, 'campaign_options', [])
        context['selected_campaign_ids'] = getattr(self, 'selected_campaign_ids', [])
        return context


class StimulusRequestExportView(LoginRequiredMixin, View):
    """Экспорт заявок на стимулирование в Excel."""

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    def get(self, request, *args, **kwargs):
        list_view = StimulusRequestListView()
        list_view.setup(request, *args, **kwargs)
        queryset = list_view.get_queryset()
        filterset = getattr(list_view, 'filterset', None)

        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"

        headers = [
            "ID",
            "Создано",
            "Обновлено",
            "Сотрудник",
            "Кампания",
            "Размер выплаты",
            "Статус",
            "Итоговый статус",
            "Ответственный",
            "Обоснование",
            "Комментарий администратора",
            "В архиве с",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment

        for row_index, request_obj in enumerate(queryset, start=2):
            created_at = request_obj.created_at.strftime('%d.%m.%Y %H:%M') if request_obj.created_at else ''
            updated_at = request_obj.updated_at.strftime('%d.%m.%Y %H:%M') if request_obj.updated_at else ''
            archived_at = request_obj.archived_at.strftime('%d.%m.%Y %H:%M') if request_obj.archived_at else ''
            responsible = request_obj.requested_by.get_full_name() or request_obj.requested_by.username
            row = [
                request_obj.pk,
                created_at,
                updated_at,
                request_obj.employee.full_name,
                request_obj.campaign.name if request_obj.campaign else '',
                float(request_obj.amount) if request_obj.amount is not None else None,
                request_obj.get_status_display(),
                request_obj.final_status or '',
                responsible,
                request_obj.justification,
                request_obj.admin_comment,
                archived_at,
            ]
            for col_index, value in enumerate(row, 1):
                ws.cell(row=row_index, column=col_index, value=value)

        self._autosize_columns(ws)

        if filterset is not None and request.GET:
            filters_sheet = wb.create_sheet("Фильтры")
            self._write_filters_sheet(filters_sheet, filterset)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"stimulus_requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _autosize_columns(self, worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value is None:
                    continue
                try:
                    cell_length = len(str(cell.value))
                except (TypeError, ValueError):
                    cell_length = 0
                if cell_length > max_length:
                    max_length = cell_length
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    def _write_filters_sheet(self, worksheet, filterset):
        worksheet.title = "Фильтры"
        worksheet.cell(row=1, column=1, value="Поле").font = self.header_font
        worksheet.cell(row=1, column=1).fill = self.header_fill
        worksheet.cell(row=1, column=1).alignment = self.header_alignment
        worksheet.cell(row=1, column=2, value="Значение").font = self.header_font
        worksheet.cell(row=1, column=2).fill = self.header_fill
        worksheet.cell(row=1, column=2).alignment = self.header_alignment

        form = filterset.form
        form_is_valid = form.is_valid()

        row_index = 2
        for field_name, field in form.fields.items():
            if form_is_valid:
                value = form.cleaned_data.get(field_name)
            else:
                value = form.data.get(field_name)

            if value in (None, '', [], (), {}):
                continue

            if isinstance(value, QuerySet):
                value = ', '.join(str(item) for item in value)
            elif isinstance(value, (list, tuple, set)):
                value = ', '.join(str(item) for item in value if item not in (None, ''))
            elif hasattr(value, 'isoformat'):
                try:
                    value = value.strftime('%d.%m.%Y')
                except (TypeError, ValueError):
                    value = str(value)
            else:
                value = str(value)

            worksheet.cell(row=row_index, column=1, value=field.label or field_name)
            worksheet.cell(row=row_index, column=2, value=value)
            row_index += 1

        self._autosize_columns(worksheet)


class StimulusRequestCreateView(LoginRequiredMixin, PermissionRequiredMixin, generic.CreateView):
    model = StimulusRequest
    form_class = StimulusRequestForm
    template_name = 'stimuli/request_form.html'
    success_url = reverse_lazy('request-list')
    permission_required = 'stimuli.add_stimulusrequest'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.requested_by = self.request.user
        response = super().form_valid(form)
        recompute_employee_totals(self.object.employee_id)
        messages.success(self.request, 'Заявка создана.')
        return response


class StimulusRequestUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = StimulusRequest
    form_class = StimulusRequestForm
    template_name = 'stimuli/request_form.html'
    success_url = reverse_lazy('request-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        instance = self.get_object()
        previous_employee_id = instance.employee_id
        response = super().form_valid(form)
        recompute_employee_totals(self.object.employee_id)
        if previous_employee_id != self.object.employee_id:
            recompute_employee_totals(previous_employee_id)
        messages.success(self.request, 'Заявка обновлена.')
        return response

    def get_queryset(self):
        base_qs = StimulusRequest.objects.select_related('employee', 'requested_by')
        user = self.request.user
        
        # Пользователи с can_view_own_requests могут редактировать только свои заявки в статусе "На рассмотрении"
        # (не могут редактировать заявки на себя, поданные другими)
        if can_view_own_requests(user):
            return base_qs.filter(requested_by=user, status=StimulusRequest.Status.PENDING)
        
        # Администраторы могут редактировать все заявки
        if can_view_all_requests(user):
            return base_qs
        
        # Руководители департамента могут редактировать заявки своего подразделения
        if is_department_manager(user):
            user_division = get_user_division(user)
            if user_division:
                return base_qs.filter(employee__division=user_division)
            return base_qs.none()
        
        # Сотрудники могут редактировать только свои заявки в статусе "На рассмотрении"
        if is_employee(user):
            return base_qs.filter(requested_by=user, status=StimulusRequest.Status.PENDING)
        
        return base_qs.none()

    def dispatch(self, request, *args, **kwargs):
        try:
            return super().dispatch(request, *args, **kwargs)
        except Http404:
            messages.error(request, 'У вас нет прав для редактирования этой заявки.')
            return redirect('request-list')


class StimulusRequestStatusUpdateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'stimuli.change_stimulusrequest'

    def get(self, request, *args, **kwargs):
        # Редиректим всех пользователей на список заявок
        return redirect('request-list')

    def post(self, request, *args, **kwargs):
        instance = get_object_or_404(StimulusRequest, pk=kwargs['pk'])
        
        # Проверяем, может ли пользователь изменять статус этой заявки
        if not can_change_request_status(request.user, instance):
            # Для сотрудников и руководителей департамента просто редиректим без сообщения об ошибке
            from .permissions import is_employee, is_department_manager
            if is_employee(request.user) or is_department_manager(request.user):
                return redirect('request-list')
            # Для остальных показываем ошибку
            messages.error(request, 'У вас нет прав на изменение статуса этой заявки.')
            return redirect('request-list')
        
        form = StimulusRequestStatusForm(request.POST, instance=instance)
        if form.is_valid():
            with transaction.atomic():
                updated_request = form.save()
                recompute_employee_totals(updated_request.employee_id)
            messages.success(request, 'Статус заявки обновлён.')
        else:
            messages.error(request, 'Не удалось обновить статус заявки. Проверьте корректность данных.')
        return redirect('request-list')


class StimulusRequestDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = StimulusRequest
    template_name = 'stimuli/request_confirm_delete.html'
    success_url = reverse_lazy('request-list')

    def get_queryset(self):
        return deletable_requests_queryset(self.request.user)

    def dispatch(self, request, *args, **kwargs):
        try:
            return super().dispatch(request, *args, **kwargs)
        except Http404:
            messages.error(request, 'У вас нет прав для удаления этой заявки.')
            return redirect('request-list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        employee_id = self.object.employee_id
        response = super().delete(request, *args, **kwargs)
        recompute_employee_totals(employee_id)
        messages.success(request, 'Заявка удалена.')
        return response


class TestView(View):
    """Простая тестовая страница"""
    
    def get(self, request, *args, **kwargs):
        import os
        return HttpResponse("""
        <html>
        <head>
            <meta charset="utf-8">
            <title>Test Page</title>
        </head>
        <body>
            <h1>✅ Django работает!</h1>
            <p>Время: """ + str(datetime.now()) + """</p>
            <p>Пользователь: """ + str(request.user) + """</p>
            <p>Авторизован: """ + str(request.user.is_authenticated) + """</p>
            <p>IP: """ + str(request.META.get('REMOTE_ADDR', 'Unknown')) + """</p>
            <p>Host: """ + str(request.META.get('HTTP_HOST', 'Unknown')) + """</p>
            <p>PORT env: """ + str(os.environ.get('PORT', 'Not set')) + """</p>
            <p>DEBUG: """ + str(os.environ.get('DJANGO_DEBUG', 'Not set')) + """</p>
            <p><a href="/">Главная</a> | <a href="/admin/">Админ</a> | <a href="/health/">Health</a></p>
        </body>
        </html>
        """, content_type='text/html; charset=utf-8')


class HomeRedirectView(View):
    """Главная страница с проверкой аутентификации"""
    
    def get(self, request, *args, **kwargs):
        # Если пользователь не авторизован, перенаправляем на страницу входа
        if not request.user.is_authenticated:
            return redirect('login')
        
        # Если авторизован, перенаправляем как раньше
        if request.user.has_perm('stimuli.view_employee'):
            return redirect('employee-list')
        return redirect('request-list')


class StimulusRequestBulkDeleteView(LoginRequiredMixin, View):
    success_url = reverse_lazy('request-list')

    def post(self, request, *args, **kwargs):
        ids = request.POST.getlist('selected_requests')
        if not ids:
            messages.warning(request, 'Не выбраны заявки для удаления.')
            return redirect(self.success_url)

        deletable_qs = deletable_requests_queryset(request.user).filter(pk__in=ids)
        deletable_ids = list(deletable_qs.values_list('pk', flat=True))

        if not deletable_ids:
            messages.error(request, 'Нет прав на удаление выбранных заявок.')
            return redirect(self.success_url)

        employees_to_update = list(deletable_qs.values_list('employee_id', flat=True))
        deleted_count, _ = deletable_qs.delete()

        for employee_id in set(employees_to_update):
            recompute_employee_totals(employee_id)

        messages.success(request, f'Удалено заявок: {deleted_count}.')
        return redirect(self.success_url)


class StimulusRequestBulkCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'stimuli.add_stimulusrequest'
    template_name = 'stimuli/request_bulk_create.html'

    def get_divisions(self):
        user = self.request.user
        
        # Администраторы видят все подразделения
        if user.is_staff:
            return Division.objects.order_by('name')
        
        # Руководители департамента видят только свое подразделение
        if is_department_manager(user):
            user_division = get_user_division(user)
            if user_division:
                return Division.objects.filter(id=user_division.id).order_by('name')
            return Division.objects.none()
        
        # Сотрудники не должны видеть подразделения в массовом создании
        return Division.objects.none()

    def get_campaigns(self):
        # Показываем только открытые кампании для массового назначения
        return RequestCampaign.objects.filter(status=RequestCampaign.Status.OPEN).order_by('-opens_at', 'name')

    def get(self, request, *args, **kwargs):
        division_id = request.GET.get('division')
        campaign_id = request.GET.get('campaign')
        
        # Для руководителей департамента автоматически выбираем их подразделение
        if not division_id and is_department_manager(request.user):
            user_division = get_user_division(request.user)
            if user_division:
                division_id = str(user_division.id)
        
        context = self._build_context(division_id, campaign_id=campaign_id)
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        division_id = request.POST.get('division')
        campaign_id = request.POST.get('campaign')
        user = request.user
        
        # Получаем доступных сотрудников в зависимости от роли пользователя
        employees_qs = get_accessible_employees(user).select_related('division').order_by('full_name')
        
        if division_id and division_id != '__all__':
            try:
                division_pk = int(division_id)
            except (TypeError, ValueError):
                messages.error(request, 'Некорректное подразделение.')
                return self.render_to_response(self._build_context(None, campaign_id=campaign_id))
            
            # Дополнительная проверка для руководителей департамента
            if is_department_manager(user):
                user_division = get_user_division(user)
                if user_division and division_pk != user_division.id:
                    messages.error(request, 'У вас нет прав для работы с этим подразделением.')
                    return self.render_to_response(self._build_context(None, campaign_id=campaign_id))
            
            employees_qs = employees_qs.filter(division_id=division_pk)
        
        employees = list(employees_qs)
        if not employees:
            messages.warning(request, 'В выбранном подразделении нет сотрудников.')
            return self.render_to_response(self._build_context(None, campaign_id=campaign_id))

        campaign = None
        if not campaign_id:
            messages.error(request, 'Необходимо выбрать кампанию.')
            return self.render_to_response(self._build_context(division_id, employees, request.POST, campaign_id=campaign_id))
        
        try:
            campaign_pk = int(campaign_id)
            campaign = RequestCampaign.objects.get(pk=campaign_pk)
            if campaign.status == RequestCampaign.Status.DRAFT:
                messages.error(request, 'Кампания в статусе "Черновик" недоступна для заявок.')
                return self.render_to_response(self._build_context(division_id, employees, request.POST, campaign_id=campaign_id))
        except (TypeError, ValueError, RequestCampaign.DoesNotExist):
            messages.error(request, 'Некорректная кампания.')
            return self.render_to_response(self._build_context(division_id, employees, request.POST, campaign_id=campaign_id))

        created = 0
        affected_employees = []
        for employee in employees:
            amount_raw = request.POST.get(f'amount_{employee.id}', '').strip()
            justification = request.POST.get(f'justification_{employee.id}', '').strip()
            if not amount_raw:
                continue
            try:
                amount = Decimal(amount_raw.replace(' ', '').replace(',', '.'))
            except Exception:
                messages.error(request, f'Некорректная сумма для {employee.full_name}.')
                return self.render_to_response(self._build_context(division_id, employees, request.POST, campaign_id=campaign_id))

            if amount <= 0:
                continue

            if not justification:
                messages.error(request, f'Обоснование обязательно для {employee.full_name}.')
                return self.render_to_response(self._build_context(division_id, employees, request.POST, campaign_id=campaign_id))

            StimulusRequest.objects.create(
                employee=employee,
                requested_by=request.user,
                amount=amount,
                justification=justification,
                campaign=campaign,
            )
            created += 1
            affected_employees.append(employee.id)

        for employee_id in set(affected_employees):
            recompute_employee_totals(employee_id)

        if created:
            messages.success(request, f'Создано заявок: {created}.')
            return redirect('request-list')

        messages.info(request, 'Не выбраны сотрудники для создания заявок.')
        return self.render_to_response(self._build_context(division_id, employees, request.POST, campaign_id=campaign_id))

    def _build_context(self, division_id=None, employees=None, data=None, campaign_id=None):
        divisions = list(self.get_divisions())
        campaigns = list(self.get_campaigns())
        employees = employees or self._get_employees_for_division(division_id)
        data = data or {}
        if isinstance(data, QueryDict):
            data = data.dict()

        if campaign_id not in (None, ''):
            selected_campaign = str(campaign_id)
        else:
            selected_campaign = str(data.get('campaign', '') or '')

        if not selected_campaign:
            default_campaign = RequestCampaign.objects.current()
            if default_campaign:
                selected_campaign = str(default_campaign.id)

        entries = []
        for employee in employees:
            entries.append({
                'employee': employee,
                'amount': data.get(f'amount_{employee.id}', ''),
                'justification': data.get(f'justification_{employee.id}', ''),
            })

        return {
            'divisions': divisions,
            'campaigns': campaigns,
            'selected_division': division_id,
            'selected_campaign': selected_campaign,
            'employee_entries': entries,
        }

    def _get_employees_for_division(self, division_id):
        user = self.request.user
        qs = get_accessible_employees(user).select_related('division').order_by('full_name')
        
        if division_id and division_id != '__all__':
            try:
                division_pk = int(division_id)
            except (TypeError, ValueError):
                return []
            
            # Дополнительная проверка для руководителей департамента
            if is_department_manager(user):
                user_division = get_user_division(user)
                if user_division and division_pk != user_division.id:
                    return []
            
            qs = qs.filter(division_id=division_pk)
        
        return list(qs)

    def render_to_response(self, context):
        from django.shortcuts import render

        return render(self.request, self.template_name, context)


def deletable_requests_queryset(user):
    base_qs = StimulusRequest.objects.select_related('employee', 'requested_by')
    
    # Пользователи с can_view_own_requests могут удалять только свои заявки в статусе "На рассмотрении"
    # (не могут удалять заявки на себя, поданные другими)
    if can_view_own_requests(user):
        return base_qs.filter(requested_by=user, status=StimulusRequest.Status.PENDING)
    
    # Администраторы могут удалять все заявки
    if can_view_all_requests(user):
        return base_qs
    
    # Руководители департамента могут удалять только свои заявки в статусе "На рассмотрении"
    if is_department_manager(user):
        return base_qs.filter(requested_by=user, status=StimulusRequest.Status.PENDING)
    
    # Сотрудники могут удалять только свои заявки в статусе "На рассмотрении"
    if is_employee(user):
        return base_qs.filter(requested_by=user, status=StimulusRequest.Status.PENDING)
    
    return base_qs.none()


class EmployeeExcelTemplateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Представление для скачивания Excel шаблона с актуальными данными сотрудников"""
    permission_required = 'stimuli.view_employee'
    logger = logging.getLogger('stimuli')

    def get(self, request, *args, **kwargs):
        self.logger.info(f"User {request.user.username} downloading Excel template")
        # Создаем новую рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"

        # Получаем все подразделения и должности для справочников
        divisions = Division.objects.all().order_by('name')
        positions = Position.objects.all().order_by('name')

        # Создаем заголовки - включаем все доступные столбцы
        headers = [
            'ФИО',
            'Подразделение',
            'Должность', 
            'Категория',
            'Ставка',
            'Оклад',
            'Выплаты по ставке',
            'Совмещения',
            'Оклад совмещений',
            'Итого базовых выплат',
            'Надбавка',
            'Основание надбавки',
            'Срок надбавки',
            'Выплата',
            'Обоснование',
            'Итого выплат'
        ]

        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Получаем всех сотрудников с их данными
        employees = Employee.objects.select_related('division', 'position').order_by('full_name')
        
        # Записываем данные сотрудников
        for row, employee in enumerate(employees, 2):
            col = 1
            ws.cell(row=row, column=col, value=employee.full_name)
            col += 1
            ws.cell(row=row, column=col, value=employee.division.name if employee.division else '')
            col += 1
            ws.cell(row=row, column=col, value=employee.position.name if employee.position else '')
            col += 1
            ws.cell(row=row, column=col, value=employee.category)
            col += 1
            ws.cell(row=row, column=col, value=float(employee.rate))
            col += 1
            ws.cell(row=row, column=col, value=float(employee.position.base_salary) if employee.position else 0.0)
            col += 1
            ws.cell(row=row, column=col, value=float(employee.salary_amount))
            col += 1
            # Совмещения - текстовое описание
            assignments_text = ''
            if employee.assignments.exists():
                assignments_list = []
                for assignment in employee.assignments.all():
                    assignment_desc = f"{assignment.position.name} ({assignment.rate})"
                    if assignment.allowance_amount:
                        assignment_desc += f" + надбавка {assignment.allowance_amount}"
                    assignments_list.append(assignment_desc)
                assignments_text = '; '.join(assignments_list)
            ws.cell(row=row, column=col, value=assignments_text)
            col += 1
            ws.cell(row=row, column=col, value=float(employee.assignments_salary_amount))
            col += 1
            ws.cell(row=row, column=col, value=float(employee.total_salary_amount))
            col += 1
            ws.cell(row=row, column=col, value=float(employee.allowance_total))
            col += 1
            ws.cell(row=row, column=col, value=employee.allowance_reason)
            col += 1
            ws.cell(row=row, column=col, value=employee.allowance_until.strftime('%d.%m.%Y') if employee.allowance_until else '')
            col += 1
            ws.cell(row=row, column=col, value=float(employee.payment))
            col += 1
            ws.cell(row=row, column=col, value=employee.justification)
            col += 1
            ws.cell(row=row, column=col, value=float(employee.total_payments))

        # Создаем лист со справочниками
        ws_div = wb.create_sheet("Подразделения")
        ws_div.cell(row=1, column=1, value="Название подразделения").font = header_font
        ws_div.cell(row=1, column=1).fill = header_fill
        
        for row, division in enumerate(divisions, 2):
            ws_div.cell(row=row, column=1, value=division.name)

        ws_pos = wb.create_sheet("Должности")
        ws_pos.cell(row=1, column=1, value="Название должности").font = header_font
        ws_pos.cell(row=1, column=2, value="Оклад").font = header_font
        ws_pos.cell(row=1, column=1).fill = header_fill
        ws_pos.cell(row=1, column=2).fill = header_fill
        
        for row, position in enumerate(positions, 2):
            ws_pos.cell(row=row, column=1, value=position.name)
            ws_pos.cell(row=row, column=2, value=float(position.base_salary))

        # Создаем лист с категориями
        ws_cat = wb.create_sheet("Категории")
        ws_cat.cell(row=1, column=1, value="Код").font = header_font
        ws_cat.cell(row=1, column=2, value="Название").font = header_font
        ws_cat.cell(row=1, column=1).fill = header_fill
        ws_cat.cell(row=1, column=2).fill = header_fill
        
        categories = Employee.Category.choices
        for row, (code, name) in enumerate(categories, 2):
            ws_cat.cell(row=row, column=1, value=str(code))
            ws_cat.cell(row=row, column=2, value=str(name))

        # Настраиваем ширину колонок
        for ws_sheet in [ws, ws_div, ws_pos, ws_cat]:
            for column in ws_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_sheet.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Создаем HTTP ответ
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"employees_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class EmployeeExcelUploadView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Представление для загрузки Excel файла с обновленными данными сотрудников"""
    permission_required = 'stimuli.add_employee'
    template_name = 'stimuli/employee_upload.html'
    logger = logging.getLogger('stimuli')

    def get(self, request, *args, **kwargs):
        form = EmployeeExcelUploadForm()
        return self.render_to_response({'form': form})

    def post(self, request, *args, **kwargs):
        form = EmployeeExcelUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return self.render_to_response({'form': form})

        excel_file = form.cleaned_data['excel_file']
        sync_mode = form.cleaned_data['sync_mode']
        
        self.logger.info(f"User {request.user.username} uploading Excel file: {excel_file.name}, sync_mode: {sync_mode}")

        try:
            from openpyxl import load_workbook
            
            # Загружаем рабочую книгу
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Получаем справочники
            divisions_dict = {d.name: d for d in Division.objects.all()}
            positions_dict = {p.name: p for p in Position.objects.all()}
            
            # Счетчики для автоматически созданных записей
            auto_created_divisions = 0
            auto_created_positions = 0
            
            # Создаем словарь категорий для обработки как кодов, так и отображаемых названий
            category_dict = {}
            for code, display_name in Employee.Category.choices:
                category_dict[code] = code  # Код категории
                category_dict[display_name] = code  # Отображаемое название
            
            created_count = 0
            updated_count = 0
            deleted_count = 0
            errors = []
            processed_employee_names = set()  # Для отслеживания обработанных сотрудников

            # Обрабатываем строки начиная со второй (пропускаем заголовки)
            for row_num in range(2, ws.max_row + 1):
                try:
                    # Читаем данные из строки (только основные поля для редактирования)
                    full_name = ws.cell(row=row_num, column=1).value
                    if not full_name or not str(full_name).strip():
                        continue
                    
                    division_name = ws.cell(row=row_num, column=2).value
                    position_name = ws.cell(row=row_num, column=3).value
                    category_code = ws.cell(row=row_num, column=4).value
                    rate = ws.cell(row=row_num, column=5).value
                    # Пропускаем столбцы 6-9 (вычисляемые поля)
                    allowance_amount = ws.cell(row=row_num, column=11).value
                    allowance_reason = ws.cell(row=row_num, column=12).value
                    allowance_until = ws.cell(row=row_num, column=13).value
                    payment = ws.cell(row=row_num, column=14).value
                    justification = ws.cell(row=row_num, column=15).value

                    # Валидация и автоматическое создание недостающих записей
                    if not division_name:
                        errors.append(f"Строка {row_num}: Не указано подразделение")
                        continue
                    
                    # Автоматически создаем подразделение если его нет
                    if division_name not in divisions_dict:
                        division_obj, created = Division.objects.get_or_create(name=division_name)
                        divisions_dict[division_name] = division_obj
                        if created:
                            auto_created_divisions += 1
                    
                    if not position_name:
                        errors.append(f"Строка {row_num}: Не указана должность")
                        continue
                    
                    # Автоматически создаем должность если её нет
                    if position_name not in positions_dict:
                        position_obj, created = Position.objects.get_or_create(
                            name=position_name,
                            defaults={'base_salary': 0}
                        )
                        positions_dict[position_name] = position_obj
                        if created:
                            auto_created_positions += 1
                    
                    if not category_code or category_code not in category_dict:
                        errors.append(f"Строка {row_num}: Неверная категория '{category_code}'")
                        continue
                    
                    # Получаем правильный код категории
                    actual_category_code = category_dict[category_code]

                    # Преобразуем числовые значения
                    try:
                        rate = float(rate) if rate is not None else 1.0
                        allowance_amount = float(allowance_amount) if allowance_amount is not None else 0.0
                        payment = float(payment) if payment is not None else 0.0
                    except (ValueError, TypeError) as e:
                        errors.append(f"Строка {row_num}: Некорректные числовые значения - {str(e)}")
                        continue

                    # Преобразуем дату
                    allowance_until_date = None
                    if allowance_until:
                        try:
                            if isinstance(allowance_until, str):
                                allowance_until_date = datetime.strptime(allowance_until, '%d.%m.%Y').date()
                            else:
                                allowance_until_date = allowance_until.date() if hasattr(allowance_until, 'date') else allowance_until
                        except ValueError:
                            errors.append(f"Строка {row_num}: Некорректный формат даты")
                            continue

                    # Проверяем, существует ли сотрудник
                    employee, created = Employee.objects.get_or_create(
                        full_name=str(full_name).strip(),
                        defaults={
                            'division': divisions_dict[division_name],
                            'position': positions_dict[position_name],
                            'category': actual_category_code,
                            'rate': rate,
                            'allowance_amount': allowance_amount,
                            'allowance_reason': str(allowance_reason).strip() if allowance_reason else '',
                            'allowance_until': allowance_until_date,
                            'payment': payment,
                            'justification': str(justification).strip() if justification else '',
                        }
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        # Обновляем существующего сотрудника
                        employee.division = divisions_dict[division_name]
                        employee.position = positions_dict[position_name]
                        employee.category = actual_category_code
                        employee.rate = rate
                        employee.allowance_amount = allowance_amount
                        employee.allowance_reason = str(allowance_reason).strip() if allowance_reason else ''
                        employee.allowance_until = allowance_until_date
                        employee.payment = payment
                        employee.justification = str(justification).strip() if justification else ''
                        employee.save()
                        updated_count += 1
                    
                    # Добавляем в список обработанных сотрудников
                    processed_employee_names.add(str(full_name).strip())

                except (ValueError, TypeError, AttributeError) as e:
                    errors.append(f"Строка {row_num}: Ошибка обработки - {str(e)}")
                    continue

            # Если выбран режим полной синхронизации, удаляем сотрудников, которых нет в файле
            if sync_mode == 'full_sync':
                all_employees = Employee.objects.all()
                for employee in all_employees:
                    if employee.full_name not in processed_employee_names:
                        employee.delete()
                        deleted_count += 1

            # Формируем сообщения для пользователя
            if auto_created_divisions > 0:
                messages.info(request, f'Автоматически создано подразделений: {auto_created_divisions}')
            if auto_created_positions > 0:
                messages.info(request, f'Автоматически создано должностей: {auto_created_positions}')
            if created_count > 0:
                messages.success(request, f'Создано новых сотрудников: {created_count}')
            if updated_count > 0:
                messages.success(request, f'Обновлено сотрудников: {updated_count}')
            if deleted_count > 0:
                messages.success(request, f'Удалено сотрудников: {deleted_count}')
            
            self.logger.info(f"Excel processing completed: created={created_count}, updated={updated_count}, deleted={deleted_count}, errors={len(errors)}")
            
            if errors:
                error_message = 'Ошибки при обработке файла:\n' + '\n'.join(errors[:10])
                if len(errors) > 10:
                    error_message += f'\n... и еще {len(errors) - 10} ошибок'
                messages.error(request, error_message)
            
            if created_count == 0 and updated_count == 0 and deleted_count == 0 and not errors:
                messages.info(request, 'Не найдено данных для обработки.')

        except (ValueError, TypeError, AttributeError, IOError) as e:
            messages.error(request, f'Ошибка при обработке файла: {str(e)}')

        return self.render_to_response({'form': form})

    def render_to_response(self, context):
        from django.shortcuts import render
        return render(self.request, self.template_name, context)

