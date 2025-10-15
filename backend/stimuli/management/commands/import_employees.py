from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from staffing.models import Division, Position
from stimuli.models import Employee


class Command(BaseCommand):
    help = 'Импортировать сотрудников из Excel-файла, экспортированного из таблицы стимулирующих выплат.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Путь к Excel-файлу (.xlsx) с данными сотрудников')
        parser.add_argument(
            '--truncate',
            action='store_true',
            help='Удалить существующих сотрудников перед импортом'
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options['xlsx_path']).expanduser()
        if not xlsx_path.exists():
            raise CommandError(f'Файл {xlsx_path} не найден')

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError('Для импорта требуется установить пакет openpyxl') from exc

        if options['truncate']:
            deleted_count, _ = Employee.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'Удалено записей сотрудников: {deleted_count}'))

        wb = load_workbook(filename=xlsx_path, read_only=True)
        ws = wb.active

        headers = []
        for cell in ws[1]:
            headers.append((cell.column_letter, (cell.value or '').strip()))

        column_map = {}
        expected = {
            'ФИО': 'full_name',
            'Подразделение': 'division',
            'Должность': 'position',
            'Категория (АУП/ППС)': 'category',
            'Выплаты': 'payment',
            'Обоснование': 'justification',
        }

        for letter, title in headers:
            if title in expected:
                column_map[expected[title]] = letter

        missing = [field for field in expected.values() if field not in column_map]
        if missing:
            raise CommandError(f'В файле отсутствуют требуемые столбцы: {", ".join(missing)}')

        created = 0
        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {headers[idx][1]: value for idx, value in enumerate(row) if idx < len(headers)}
            full_name = (row_data.get('ФИО') or '').strip()

            if not full_name or full_name.upper() == 'ВСЕГО':
                skipped += 1
                continue

            division_name = (row_data.get('Подразделение') or '').strip() or 'Не указано'
            division_obj, _ = Division.objects.get_or_create(name=division_name)
            position_name = (row_data.get('Должность') or '').strip() or 'Без должности'
            position_obj, _ = Position.objects.get_or_create(name=position_name)
            category = (row_data.get('Категория (АУП/ППС)') or '').strip() or Employee.Category.OTHER

            if category not in dict(Employee.Category.choices):
                category = Employee.Category.OTHER

            payment_value = row_data.get('Выплаты')
            payment = Decimal('0')
            if payment_value not in (None, ''):
                try:
                    payment = Decimal(str(payment_value).replace(' ', '').replace(',', '.'))
                except Exception:
                    self.stdout.write(self.style.WARNING(
                        f'Не удалось преобразовать сумму "{payment_value}" для {full_name}. Установлен 0.'
                    ))

            justification = (row_data.get('Обоснование') or '').strip()

            obj, created_flag = Employee.objects.update_or_create(
                full_name=full_name,
                defaults={
                    'division': division_obj,
                    'position': position_obj,
                    'category': category,
                    'rate': 1,
                    'allowance_amount': Decimal('0'),
                    'allowance_reason': '',
                    'payment': payment,
                    'justification': justification,
                },
            )

            if created_flag:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f'Импорт завершён. Создано: {created}, обновлено: {updated}, пропущено: {skipped}'
        ))
        wb.close()
