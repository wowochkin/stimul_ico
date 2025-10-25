from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from staffing.models import Division, Position


class Command(BaseCommand):
    help = 'Создает подразделения и должности из Excel файла с сотрудниками'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Путь к Excel-файлу (.xlsx) с данными сотрудников')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Показать что будет создано без фактического создания'
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options['xlsx_path']).expanduser()
        if not xlsx_path.exists():
            raise CommandError(f'Файл {xlsx_path} не найден')

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError('Для импорта требуется установить пакет openpyxl') from exc

        dry_run = options['dry_run']
        
        wb = load_workbook(filename=xlsx_path, read_only=True)
        ws = wb.active

        # Получаем заголовки
        headers = []
        for cell in ws[1]:
            headers.append((cell.column_letter, (cell.value or '').strip()))

        # Находим колонки с подразделениями и должностями
        division_col = None
        position_col = None
        
        for letter, title in headers:
            if title == 'Подразделение':
                division_col = letter
            elif title == 'Должность':
                position_col = letter

        if not division_col:
            raise CommandError('В файле не найдена колонка "Подразделение"')
        if not position_col:
            raise CommandError('В файле не найдена колонка "Должность"')

        # Собираем уникальные подразделения и должности
        divisions = set()
        positions = set()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {headers[idx][1]: value for idx, value in enumerate(row) if idx < len(headers)}
            
            division_name = (row_data.get('Подразделение') or '').strip()
            position_name = (row_data.get('Должность') or '').strip()
            
            if division_name and division_name.upper() != 'ВСЕГО':
                divisions.add(division_name)
            
            if position_name and position_name.upper() != 'ВСЕГО':
                positions.add(position_name)

        wb.close()

        # Создаем подразделения
        created_divisions = []
        for division_name in sorted(divisions):
            if dry_run:
                if not Division.objects.filter(name=division_name).exists():
                    created_divisions.append(division_name)
                    self.stdout.write(f'[DRY RUN] Будет создано подразделение: {division_name}')
                else:
                    self.stdout.write(f'Подразделение уже существует: {division_name}')
            else:
                division_obj, created = Division.objects.get_or_create(name=division_name)  # noqa: E501
                if created:
                    created_divisions.append(division_name)
                    self.stdout.write(self.style.SUCCESS(f'✓ Создано подразделение: {division_name}'))  # noqa: E501
                else:
                    self.stdout.write(f'Подразделение уже существует: {division_name}')

        # Создаем должности
        created_positions = []
        for position_name in sorted(positions):
            if dry_run:
                if not Position.objects.filter(name=position_name).exists():
                    created_positions.append(position_name)
                    self.stdout.write(f'[DRY RUN] Будет создана должность: {position_name}')
                else:
                    self.stdout.write(f'Должность уже существует: {position_name}')
            else:
                position_obj, created = Position.objects.get_or_create(  # noqa: E501
                    name=position_name,
                    defaults={'base_salary': 0}
                )
                if created:
                    created_positions.append(position_name)
                    self.stdout.write(self.style.SUCCESS(f'✓ Создана должность: {position_name}'))  # noqa: E501
                else:
                    self.stdout.write(f'Должность уже существует: {position_name}')

        # Итоговая статистика
        if dry_run:
            self.stdout.write(self.style.WARNING(f'\n[DRY RUN] Результат:'))
            self.stdout.write(f'Будет создано подразделений: {len(created_divisions)}')
            self.stdout.write(f'Будет создано должностей: {len(created_positions)}')
        else:
            self.stdout.write(self.style.SUCCESS(f'\n✓ Обработка завершена:'))
            self.stdout.write(f'Создано подразделений: {len(created_divisions)}')
            self.stdout.write(f'Создано должностей: {len(created_positions)}')
            
            if created_divisions:
                self.stdout.write(f'\nСозданные подразделения:')
                for div in created_divisions:
                    self.stdout.write(f'  - {div}')
            
            if created_positions:
                self.stdout.write(f'\nСозданные должности:')
                for pos in created_positions:
                    self.stdout.write(f'  - {pos}')
