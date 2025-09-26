from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import transaction
from django.db.models import Prefetch
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import generic, View

from .forms import PositionQuotaForm, PositionQuotaVersionForm
from .models import Division, PositionQuota, PositionQuotaVersion


class PositionQuotaListView(LoginRequiredMixin, PermissionRequiredMixin, generic.TemplateView):
    template_name = 'staffing/position_quota_list.html'
    permission_required = 'staffing.view_positionquota'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        version_qs = PositionQuotaVersion.objects.order_by('-effective_from', '-created_at')
        quota_qs = PositionQuota.objects.select_related('position').prefetch_related(
            Prefetch('versions', queryset=version_qs)
        ).order_by('position__name')
        divisions = Division.objects.prefetch_related(
            Prefetch('quotas', queryset=quota_qs)
        ).order_by('name')
        today = timezone.now().date()
        for division in divisions:
            for quota in division.quotas.all():
                quota.version_form = PositionQuotaVersionForm(initial={
                    'effective_from': today,
                    'total_fte': quota.total_fte,
                    'occupied_fte': quota.occupied_fte,
                    'vacant_fte': quota.vacant_fte,
                })
        context['divisions'] = divisions
        context['create_form'] = PositionQuotaForm()
        context['export_url'] = reverse('staffing:quota-export')
        return context


class PositionQuotaCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'staffing.add_positionquota'

    def post(self, request, *args, **kwargs):
        form = PositionQuotaForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                quota = form.save()
                PositionQuotaVersion.objects.create(
                    quota=quota,
                    effective_from=timezone.now().date(),
                    total_fte=quota.total_fte,
                    occupied_fte=quota.occupied_fte,
                    vacant_fte=quota.vacant_fte,
                )
            messages.success(request, 'Позиция добавлена в штатное расписание.')
        else:
            messages.error(request, 'Не удалось создать позицию. Проверьте введённые данные.')
        return redirect('staffing:quota-list')


class PositionQuotaVersionCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'staffing.change_positionquota'

    def post(self, request, pk, *args, **kwargs):
        quota = get_object_or_404(PositionQuota, pk=pk)
        form = PositionQuotaVersionForm(request.POST)
        if form.is_valid():
            version = form.save(commit=False)
            version.quota = quota
            if version.effective_to and version.effective_to < version.effective_from:
                form.add_error('effective_to', 'Дата окончания не может быть раньше даты начала.')
                messages.error(request, 'Дата окончания не может быть раньше даты начала.')
            else:
                version.save()
                quota.total_fte = version.total_fte
                quota.occupied_fte = version.occupied_fte
                quota.vacant_fte = version.vacant_fte
                quota.save(update_fields=['total_fte', 'occupied_fte', 'vacant_fte', 'updated_at'])
                messages.success(request, 'Изменения ставок сохранены.')
                return redirect('staffing:quota-list')
        else:
            messages.error(request, 'Не удалось обновить ставки. Проверьте введённые данные.')
        return redirect('staffing:quota-list')


class PositionQuotaUpdateView(LoginRequiredMixin, PermissionRequiredMixin, generic.UpdateView):
    model = PositionQuota
    form_class = PositionQuotaForm
    template_name = 'staffing/position_quota_form.html'
    permission_required = 'staffing.change_positionquota'
    success_url = reverse_lazy('staffing:quota-list')

    def form_valid(self, form):
        original = PositionQuota.objects.get(pk=form.instance.pk)
        response = super().form_valid(form)
        updated = self.object
        changed = any([
            original.total_fte != updated.total_fte,
            original.occupied_fte != updated.occupied_fte,
            original.vacant_fte != updated.vacant_fte,
        ])
        if changed:
            PositionQuotaVersion.objects.create(
                quota=updated,
                effective_from=timezone.now().date(),
                total_fte=updated.total_fte,
                occupied_fte=updated.occupied_fte,
                vacant_fte=updated.vacant_fte,
            )
        messages.success(self.request, 'Позиция обновлена.')
        return response


class PositionQuotaDeleteView(LoginRequiredMixin, PermissionRequiredMixin, generic.DeleteView):
    model = PositionQuota
    template_name = 'staffing/position_quota_confirm_delete.html'
    permission_required = 'staffing.delete_positionquota'
    success_url = reverse_lazy('staffing:quota-list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        quota_name = str(self.object)
        messages.success(request, f'Позиция "{quota_name}" удалена.')
        return super().delete(request, *args, **kwargs)


class PositionQuotaExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'staffing.view_positionquota'

    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Штатное расписание'
        header = ['Подразделение', 'Должность', 'Всего ставок', 'Занятые', 'Вакантные', 'Комментарий', 'Дата актуальности']
        sheet.append(header)

        version_prefetch = Prefetch(
            'versions',
            queryset=PositionQuotaVersion.objects.order_by('-effective_from', '-created_at')
        )
        quotas_qs = PositionQuota.objects.select_related('position', 'division').prefetch_related(version_prefetch).order_by('division__name', 'position__name')

        for quota in quotas_qs:
            versions = list(quota.versions.all())
            latest_version = versions[0] if versions else None
            effective_from = latest_version.effective_from if latest_version else None
            sheet.append([
                quota.division.name,
                quota.position.name,
                float(quota.total_fte),
                float(quota.occupied_fte),
                float(quota.vacant_fte),
                quota.comment or '',
                effective_from.strftime('%d.%m.%Y') if effective_from else '',
            ])

        for idx, column in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            sheet.column_dimensions[get_column_letter(idx)].width = max(15, max_length + 2)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename="position_quota_{timestamp}.xlsx"'
        workbook.save(response)
        return response
